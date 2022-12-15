import openpyxl
from json.tool import main
import os
import pandas as pd
# Set directory
directory = '/Users/kirdemir/desktop'
files = os.listdir(directory)
mainList = []
listOfNumbers = []
f = open('errors.txt','w')
# Set main .xlsx file path. Read values and put them in a list.
file_path = '/Users/kirdemir/desktop/Yeni/Untitled 3.xlsx'
ab = pd.read_excel(os.path.join(file_path))
for i in ab[210202017]:
    mainList.append(i)
srcfile = openpyxl.load_workbook(
    '/Users/kirdemir/desktop/Yeni/Untitled 3.xlsx', read_only=False, keep_vba=True)
for file in files:
    # Open all .csv files
    if file.endswith('.csv'):
        df = pd.read_csv(os.path.join(directory, file))

        if len(df.columns) > 3:

            for i in range(len(df['Unnamed: 3'])):
                try:
                    # Take values and put them in another list
                    if (len(df['Unnamed: 3'][i]) > 5 and len(df['Unnamed: 3'][i]) < 12 and str(df['Unnamed: 3'][i])[:2].isnumeric() and int(df['Unnamed: 3'][i][:2]) >= 30) or len(df['Unnamed: 3']) > 15:
                        try:
                            listOfNumbers.append(str(df['Unnamed: 4'][i])[:9])
                        except:
                            f.write(df['Unnamed: 4'][i] + " listeye eklenemedi.\n")
                            continue
                except:
                    continue
# Compare listed values and write them in certain cells in new .xlsx file
    for i in range(len(listOfNumbers)):
        if listOfNumbers[i].isnumeric() and int(listOfNumbers[i]) in mainList:
            sheetname = srcfile.get_sheet_by_name(srcfile.sheetnames[0])
            if sheetname.cell(row=mainList.index(
                    int(listOfNumbers[i]))+2, column=5).value is None:
                sheetname.cell(row=mainList.index(
                    int(listOfNumbers[i]))+2, column=5).value = 'X'
            elif sheetname.cell(row=mainList.index(
                    int(listOfNumbers[i]))+2, column=6).value is None:
                sheetname.cell(row=mainList.index(
                    int(listOfNumbers[i]))+2, column=6).value = 'X'
            elif sheetname.cell(row=mainList.index(
                    int(listOfNumbers[i]))+2, column=7).value is None:
                sheetname.cell(row=mainList.index(
                    int(listOfNumbers[i]))+2, column=7).value = 'X'
            elif sheetname.cell(row=mainList.index(
                    int(listOfNumbers[i]))+2, column=8).value is None:
                sheetname.cell(row=mainList.index(
                    int(listOfNumbers[i]))+2, column=8).value = 'X'
            else:
                continue
    listOfNumbers = []
srcfile.save('new file.xlsx')
